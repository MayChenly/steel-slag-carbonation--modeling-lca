"""Python port of MATLAB LCA pipeline for steel slag carbonation.

Direct conversion of:
  - BOF_country_lca.m / BF_country_lca.m / EAF_country_lca.m  (37-country loop)
  - BOF_lca.m / BF_lca.m / EAF_lca.m                          (single-country UK demo)

Uses OWID 2025 grid carbon intensity values.
Bug fix: UK CI now 0.21741 (was hardcoded 0.2376 in single-country .m files).
"""
import math
import csv
import os
from pathlib import Path

import openpyxl

DATA_ROOT = Path('/sessions/gallant-loving-edison/mnt/LCApart')
OUT_ROOT = Path('/sessions/gallant-loving-edison/mnt/outputs/lca_python')
OUT_ROOT.mkdir(parents=True, exist_ok=True)

# ─── OWID 2025 emission factors (kg CO2 / kWh) ───
OWID_2025_CI = {
    'Austria': 0.11691, 'Belgium': 0.14982, 'Bulgaria': 0.27555, 'Croatia': 0.15850,
    'Czechia': 0.40146, 'Finland': 0.05747, 'France': 0.04144, 'Germany': 0.32965,
    'Greece': 0.31509, 'Hungary': 0.16302, 'Italy': 0.28478, 'Luxembourg': 0.12338,
    'Netherlands': 0.25356, 'Poland': 0.58860, 'Portugal': 0.12791, 'Romania': 0.25075,
    'Slovakia': 0.09485, 'Slovenia': 0.18329, 'Spain': 0.15360, 'Sweden': 0.03526,
    'Türkiye': 0.47474, 'United Kingdom': 0.21741, 'Russia': 0.44973,
    'Canada': 0.19072, 'Mexico': 0.47402, 'United States': 0.38440,
    'Argentina': 0.34600, 'Brazil': 0.10995, 'Chile': 0.28949,
    'Egypt': 0.56323, 'South Africa': 0.69929, 'Iran': 0.65953,
    'China': 0.52534, 'India': 0.67013, 'Japan': 0.47726,
    'South Korea': 0.41706, 'Australia': 0.52518,
}

# ─── Fixed parameters (same for BOF / BF / EAF; only D_sm differs) ───
V_L = 2000.0                  # L liquid in precipitation reactor
CO2_MOLAR_MASS = 44.0         # g/mol
W_I = 0.0135                  # Bond Work Index (kWh/kg)
DP_FINAL = 1e-5               # final particle diameter (m, = 10 μm)

# Reactor 1 (dissolution) constant power (W)
E_REACTOR1_W = 99.8121

# Reactor 2 (precipitation) parameters
RHO_L = 1050.0     # kg/m³
MU_L = 1.2e-3      # Pa·s
N_RPS = 1.5        # rps
D = 0.6            # impeller diameter (m)
DI = 0.12          # impeller blade width (m)
VL_M3 = 2.0        # liquid volume (m³)
G = 9.81           # gravity (m/s²)

# Fan parameters
H_L = 1.0          # liquid column height (m)
TOTAL_PRESSURE_BAR = 1.0
ETA = 0.85 * 0.95
GAMMA = 1.28

# DAC option 2 (low-purity, Climeworks)
ELECTRICITY_PER_KGCO2_010BAR = 4552.583826 / 3.6 / 1000  # kWh per kg CO2 (was /3.6 for /tonne, /1000 for /kg)
# So actually let me recompute carefully:
# Original: electrcity_for_010_bar_CO2 = 4552.583826 / 3.6 = 1264.6 kWh/tonne_CO2
# Then E_DAC = electricity_for_010_bar_CO2/1000 * CO2_supplied_kg = 1.2646 * CO2_kg
ELECTRICITY_PER_TONNE_CO2_010BAR = 4552.583826 / 3.6  # kWh/tonne CO2

# BPMED energy
E_BPMED_PER_2M3_VOLUME_J = 17328.25801  # baseline value
# E_BPMED_system = 17328.25801 * (V_L/2) / 3.6e6  # kWh per batch
E_BPMED_PER_BATCH_KWH = 17328.25801 * (V_L / 2) / 3.6e6  # ~2.4 kWh/batch

# Initial slag particle diameter (raw material, before grinding) — varies by slag type
# BOF: 15 mm = 0.015 m (initial size); also BOF_D_sm = 0.0025 in country script (!! discrepancy)
# Inspect BF and EAF to confirm
SLAG_INITIAL_D = {
    'BOF': 0.015,   # Note: country script uses 0.0025; single script uses 0.015. Check!
    'BF':  0.0025,  # placeholder; will verify
    'EAF': 0.0095,  # placeholder; will verify
}

# ─── LCA computation ───
def compute_energies(slag_kg, dissolution_min, caco3_conc_molperL,
                     precipitation_h, q_feed_m3perh, D_sm):
    """Compute per-batch energy demands for each unit operation (kWh)."""
    # Crushing energy (Bond's law, factor 0.01 already in MATLAB; result in kWh per batch)
    E_crushing = 0.01 * W_I * (1.0 / math.sqrt(DP_FINAL) - 1.0 / math.sqrt(D_sm)) * slag_kg

    # Reactor 1 stirring energy
    E_stirring_R1 = (E_REACTOR1_W / 1000.0) * (dissolution_min / 60.0)

    # Reactor 2 stirring energy (Reynolds-Froude unbaffled correlation)
    Q_G = q_feed_m3perh / 3600.0  # m³/s
    Re_L = (RHO_L * N_RPS * D**2) / MU_L
    Fr_L = (D * N_RPS**2) / G
    Np_1 = 19.5 * Re_L**(-0.3)
    Np_2 = 24.0 * (Re_L * Fr_L)**(-1.0/3.0)
    Np = min(Np_1, Np_2)
    P_0 = RHO_L * N_RPS**3 * D**5 * Np
    term1 = (Q_G / (N_RPS * VL_M3))**(-0.25)
    term2 = ((N_RPS**2 * D**4) / (G * DI * VL_M3**(2.0/3.0)))**(-0.2)
    PG_P0 = 0.1 * term1 * term2
    PG = P_0 * PG_P0  # W
    E_stirring_R2 = (PG / 1000.0) * precipitation_h  # kWh

    # Fan energy
    delta_P = RHO_L * G * H_L     # Pa
    P_in = TOTAL_PRESSURE_BAR * 1e5  # Pa
    P_out = P_in + delta_P
    E_fan_W = (1.0/ETA) * P_in * Q_G * GAMMA / (GAMMA - 1) * \
              ((P_out / P_in)**(1 - 1/GAMMA) - 1)
    E_fan = (E_fan_W / 1000.0) * precipitation_h  # kWh

    # CO2 stored (mass-based)
    CO2_stored_kg = caco3_conc_molperL * V_L * CO2_MOLAR_MASS / 1000.0

    # DAC energy (electricity for delivering CO2 at 0.1 bar)
    E_DAC = ELECTRICITY_PER_TONNE_CO2_010BAR / 1000.0 * CO2_stored_kg  # kWh

    return {
        'E_crushing': E_crushing,
        'E_stirring_R1': E_stirring_R1,
        'E_stirring_R2': E_stirring_R2,
        'E_fan': E_fan,
        'E_DAC': E_DAC,
        'E_BPMED': E_BPMED_PER_BATCH_KWH,
        'CO2_stored_kg': CO2_stored_kg,
    }

def compute_country_lca(slag_type, slag_kg, dissolution_min, caco3_conc, prec_h, q_feed, D_sm):
    """Run the 37-country LCA loop for one slag type."""
    energies = compute_energies(slag_kg, dissolution_min, caco3_conc, prec_h, q_feed, D_sm)

    E_total = (energies['E_DAC'] + energies['E_BPMED'] +
               energies['E_stirring_R1'] + energies['E_stirring_R2'] +
               energies['E_fan'] + energies['E_crushing'])

    CO2_stored_kg = energies['CO2_stored_kg']
    conversion_ratio = 1000.0 / slag_kg  # scaling factor to per-tonne

    rows = []
    rows.append(['Country', 'Emission factor', 'DAC', 'BPMED system',
                 'Stirring dissolution', 'Stirring precipitation',
                 'Fan energy', 'Crushing', 'Total emission',
                 'Net emission', 'Net emission per tonne'])

    for country, ci in sorted(OWID_2025_CI.items()):
        dac_em = energies['E_DAC'] * ci
        bpmed_em = energies['E_BPMED'] * ci
        stir1_em = energies['E_stirring_R1'] * ci
        stir2_em = energies['E_stirring_R2'] * ci
        fan_em = energies['E_fan'] * ci
        crush_em = energies['E_crushing'] * ci
        total_em = E_total * ci
        net_em = total_em - CO2_stored_kg
        net_em_tonne = net_em * conversion_ratio
        rows.append([country, ci, dac_em, bpmed_em, stir1_em, stir2_em,
                     fan_em, crush_em, total_em, net_em, net_em_tonne])

    return rows, energies, E_total, CO2_stored_kg

def read_dissolution_input(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['Sheet1']
    row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    return {
        'slag_kg': row[0],
        'dissolution_min': row[1],
        'caco3_conc': row[3],
        'prec_h': row[4],
        'q_feed': row[5],
    }

def write_xlsx(rows, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    wb.save(out_path)

# ─── Run for each slag type ───
SLAG_INPUT_FILES = {
    'BOF': DATA_ROOT / 'BOF/BOF_Indirect_dissolution_result.xlsx',
    'BF':  DATA_ROOT / 'BF/BF_Indirect_dissolution_result.xlsx',
    'EAF': DATA_ROOT / 'EAF/EAF_Indirect_dissolution_result.xlsx',
}

# D_sm (initial size for crushing) - need to read original .m files to confirm
# From MATLAB: BOF_country_lca.m uses BOF_D_sm = 0.0025 (this looks wrong vs single)
# Let me use the same value as country_lca.m for consistency
SLAG_D_SM = {
    'BOF': 0.0025,   # matching BOF_country_lca.m line 27
    'BF':  0.0025,   # placeholder, will verify
    'EAF': 0.0025,   # placeholder, will verify
}

print('═' * 80)
print('Python LCA pipeline — OWID 2025 CI values')
print('═' * 80)

for slag_type, input_path in SLAG_INPUT_FILES.items():
    if not input_path.exists():
        print(f'\n[{slag_type}] input file not found, skipping')
        continue
    print(f'\n[{slag_type}] reading {input_path.name}…')
    inp = read_dissolution_input(input_path)
    print(f'  input: slag={inp["slag_kg"]} kg, diss={inp["dissolution_min"]} min, '
          f'caco3={inp["caco3_conc"]:.5f} mol/L, prec={inp["prec_h"]} h, qfeed={inp["q_feed"]:.2f} m³/h')

    D_sm = SLAG_D_SM[slag_type]
    rows, energies, E_total, CO2_stored = compute_country_lca(
        slag_type, inp['slag_kg'], inp['dissolution_min'], inp['caco3_conc'],
        inp['prec_h'], inp['q_feed'], D_sm
    )

    print(f'  E_DAC = {energies["E_DAC"]:.4f} kWh/batch')
    print(f'  E_BPMED = {energies["E_BPMED"]:.4f} kWh/batch')
    print(f'  E_crushing = {energies["E_crushing"]:.4f} kWh/batch')
    print(f'  E_stirring_R1 = {energies["E_stirring_R1"]:.4f} kWh/batch')
    print(f'  E_stirring_R2 = {energies["E_stirring_R2"]:.4f} kWh/batch')
    print(f'  E_fan = {energies["E_fan"]:.4f} kWh/batch')
    print(f'  E_total = {E_total:.4f} kWh/batch')
    print(f'  CO2_stored = {CO2_stored:.4f} kg/batch')

    out_path = OUT_ROOT / f'{slag_type}_energy_output_by_country.xlsx'
    write_xlsx(rows, out_path)
    print(f'  ✓ wrote {out_path.name}')

    # Show UK row
    uk_row = next((r for r in rows if r[0] == 'United Kingdom'), None)
    if uk_row:
        print(f'\n  UK row preview:')
        print(f'    Emission factor: {uk_row[1]:.5f}')
        print(f'    Total emission: {uk_row[8]:.4f} kg/batch')
        print(f'    Net emission: {uk_row[9]:.4f} kg/batch')
        print(f'    Net emission per tonne: {uk_row[10]:.2f} kg/t')

print('\nAll outputs in:', OUT_ROOT)
