"""Build the complete LCA dataset for upload — ALL scenarios × ALL countries × ALL slag types.

Output one master XLSX with multiple sheets covering:
  - 37 OWID 2025 country CIs
  - Direct (gas-solid) LCA: 4 scenarios × 7 distances × 37 countries × 3 slag types = 3,108 rows
    with full component breakdown (DAC, liquefaction, vaporisation, transport, crushing, fan, escape, DAC capture, R_total)
  - Indirect aqueous LCA: 37 countries × 3 slag types = 111 rows
    with full breakdown (DAC 0.1 bar, BPMED, stir R1, stir R2, fan, crushing, total, net)
  - Master per-tonne summary: best pathway per country × slag type
  - National CDR potential: best per-tonne × slag production = kt CO2 / yr
"""
import math
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

PROJECT_ROOT = Path('/sessions/gallant-loving-edison/mnt/Project3_Steel slag/direct carbonation/results/LCApart')
INDIRECT_INPUT_ROOT = Path('/sessions/gallant-loving-edison/mnt/LCApart')
OUT_PATH = Path('/sessions/gallant-loving-edison/mnt/new thesis/Complete_LCA_OWID2025.xlsx')

# ─── OWID 2025 CI ───
OWID_2025 = {
    'Austria': 0.11691, 'Belgium': 0.14982, 'Bulgaria': 0.27555, 'Croatia': 0.15850,
    'Czechia': 0.40146, 'Finland': 0.05747, 'France': 0.04144, 'Germany': 0.32965,
    'Greece': 0.31509, 'Hungary': 0.16302, 'Italy': 0.28478, 'Luxembourg': 0.12338,
    'Netherlands': 0.25356, 'Poland': 0.58860, 'Portugal': 0.12791, 'Romania': 0.25075,
    'Slovakia': 0.09485, 'Slovenia': 0.18329, 'Spain': 0.15360, 'Sweden': 0.03526,
    'Türkiye': 0.47474, 'United Kingdom': 0.21741, 'Russia': 0.44973,
    'Canada': 0.19072, 'Mexico': 0.47402, 'United States': 0.38440,
    'Argentina': 0.34600, 'Brazil': 0.10995, 'Chile': 0.28949,
    'Egypt': 0.56323, 'South Africa': 0.69929, 'Iran': 0.65953,
    'China': 0.52534, 'India': 0.67013, 'Japan': 0.47726,
    'South Korea': 0.41706, 'Australia': 0.52518,
}

# National slag production (Mt/yr) — averages from Table S8 / Appendix C-2
SLAG_PROD = {
    'Austria':       {'BOF': 0.8531, 'EAF': 0.0844, 'BF': 1.5125},
    'Belgium':       {'BOF': 0.6064, 'EAF': 0.2686, 'BF': 0.9900},
    'Bulgaria':      {'BOF': 0.0000, 'EAF': 0.0625, 'BF': 0.0000},
    'Croatia':       {'BOF': 0.0000, 'EAF': 0.0250, 'BF': 0.0000},
    'Czechia':       {'BOF': 0.5165, 'EAF': 0.0210, 'BF': 0.7425},
    'Finland':       {'BOF': 0.2472, 'EAF': 0.1903, 'BF': 0.6050},
    'France':        {'BOF': 1.0194, 'EAF': 0.4931, 'BF': 1.7050},
    'Germany':       {'BOF': 3.2292, 'EAF': 1.3708, 'BF': 6.4900},
    'Greece':        {'BOF': 0.0000, 'EAF': 0.1875, 'BF': 0.0000},
    'Hungary':       {'BOF': 0.0775, 'EAF': 0.0350, 'BF': 0.0550},
    'Italy':         {'BOF': 0.4320, 'EAF': 2.2680, 'BF': 0.8525},
    'Luxembourg':    {'BOF': 0.0000, 'EAF': 0.2375, 'BF': 0.0000},
    'Netherlands':   {'BOF': 0.4781, 'EAF': 0.2844, 'BF': 1.1275},
    'Poland':        {'BOF': 0.4311, 'EAF': 0.4940, 'BF': 0.7425},
    'Portugal':      {'BOF': 0.0000, 'EAF': 0.2375, 'BF': 0.0000},
    'Romania':       {'BOF': 0.2038, 'EAF': 0.1212, 'BF': 0.1650},
    'Slovakia':      {'BOF': 0.3846, 'EAF': 0.1029, 'BF': 0.0000},
    'Slovenia':      {'BOF': 0.0000, 'EAF': 0.0750, 'BF': 0.0000},
    'Spain':         {'BOF': 0.4600, 'EAF': 0.9775, 'BF': 0.8250},
    'Sweden':        {'BOF': 0.3586, 'EAF': 0.1914, 'BF': 0.7975},
    'Türkiye':       {'BOF': 1.2504, 'EAF': 3.1371, 'BF': 2.3925},
    'United Kingdom':{'BOF': 0.6068, 'EAF': 0.1433, 'BF': 1.2375},
    'Russia':        {'BOF': 5.8094, 'EAF': 2.9583, 'BF': 15.0150},
    'Canada':        {'BOF': 0.8152, 'EAF': 0.6973, 'BF': 1.5950},
    'Mexico':        {'BOF': 0.3303, 'EAF': 1.9322, 'BF': 0.2750},
    'United States': {'BOF': 3.1194, 'EAF': 6.9431, 'BF': 5.6650},
    'Argentina':     {'BOF': 0.2843, 'EAF': 0.3532, 'BF': 0.5500},
    'Brazil':        {'BOF': 3.2011, 'EAF': 1.0145, 'BF': 7.0675},
    'Chile':         {'BOF': 0.0914, 'EAF': 0.0587, 'BF': 0.1650},
    'Egypt':         {'BOF': 0.0000, 'EAF': 0.8244, 'BF': 0.0000},
    'South Africa':  {'BOF': 0.3630, 'EAF': 0.1870, 'BF': 0.7975},
    'Iran':          {'BOF': 0.3137, 'EAF': 3.5114, 'BF': 0.9625},
    'China':         {'BOF': 116.7135,'EAF': 13.1366,'BF': 242.8800},
    'India':         {'BOF': 7.1734, 'EAF': 8.4891, 'BF': 23.7325},
    'Japan':         {'BOF': 8.1730, 'EAF': 2.9771, 'BF': 17.3250},
    'South Korea':   {'BOF': 5.6341, 'EAF': 2.5909, 'BF': 12.4300},
    'Australia':     {'BOF': 0.5237, 'EAF': 0.1888, 'BF': 0.9625},
}

DISTANCES = list(range(0, 3001, 500))  # 7 distances

# ─── DAC / transport parameters ───
TOTAL_WEIGHT = 1.0 + 2 * 0.4212
ROAD_EF = 0.112
EV_KWH_PER_KM = 0.921

O1_CO2_AIR = 0.769231
O1_ELEC = 152.8836
O2_CO2_AIR = 1.0
O2_ELEC = 7220.0 / 3.6

LIQUEFACTION = 89.0
VAPORISATION = 5.4

# ─── Load gas-solid per-slag-type LCA input ───
def load_slag_lca_data(slag_type):
    fp = PROJECT_ROOT / f'{slag_type}_lca_data.xlsx'
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    def col(name):
        return headers.index(name) + 1
    row1 = {h: ws.cell(2, col(h)).value for h in headers if h is not None}
    return row1

# ─── Direct (gas-solid) LCA computation ───
def direct_lca_per_country(slag_data, ci, distance):
    """Return dict with all 4 scenarios and all component breakdowns at per-tonne basis."""
    mass = slag_data['SS_total_mass_kg']
    fan_MJ = slag_data['fan_energy_MJ']
    crush_kWh = slag_data['crushing_energy_kWh']
    CO2_sup = slag_data['CO2_total_supplied_kg']
    CO2_str = slag_data['CO2_total_stored_kg']
    conv = 1000.0 / mass

    input_ratio = CO2_sup / 1000.0 * conv

    # Per-tonne components (CI-dependent)
    E_crushing = crush_kWh * ci * conv               # kg CO2/t slag
    E_fan = fan_MJ / 3.6 * ci * conv
    E_escape = (CO2_sup - CO2_str) * conv            # kg/t (mass)
    CO2_stored_pt = CO2_str * conv

    # DAC processing (per tonne CO2 supplied, then scale by input_ratio)
    o1_E_DAC = O1_ELEC * ci
    o2_E_DAC = O2_ELEC * ci
    E_liq = LIQUEFACTION * ci
    E_vap = VAPORISATION * ci

    # DAC capture credits (mass-based)
    o1_R_DAC = CO2_sup * O1_CO2_AIR * conv
    o2_R_DAC = CO2_sup * O2_CO2_AIR * conv

    # Transport
    E_tr_diesel = TOTAL_WEIGHT * distance * 2 * ROAD_EF  # kg CO2/t CO2 supplied
    E_tr_EV = distance * 2 * EV_KWH_PER_KM * ci

    out = {}
    for trans_name, trans_kg in [('Diesel', E_tr_diesel), ('EV', E_tr_EV)]:
        for opt, opt_DAC, opt_RDAC in [('Opt1', o1_E_DAC, o1_R_DAC), ('Opt2', o2_E_DAC, o2_R_DAC)]:
            R_part1 = opt_DAC + E_liq + E_vap + trans_kg
            E_part1 = R_part1 * input_ratio
            E_DAC_pt = opt_DAC * input_ratio
            E_liq_pt = E_liq * input_ratio
            E_vap_pt = E_vap * input_ratio
            E_tr_pt = trans_kg * input_ratio
            total_em = E_part1 + E_crushing + E_fan + E_escape
            R_total = total_em - opt_RDAC
            out[f'{trans_name}+{opt}'] = {
                'DAC_process_kg': E_DAC_pt,
                'Liquefaction_kg': E_liq_pt,
                'Vaporisation_kg': E_vap_pt,
                'Transport_kg': E_tr_pt,
                'Crushing_kg': E_crushing,
                'Fan_kg': E_fan,
                'CO2_escape_kg': E_escape,
                'Total_emission_kg': total_em,
                'DAC_capture_kg': -opt_RDAC,  # negative = credit
                'CO2_stored_kg': CO2_stored_pt,
                'Net_emission_kg_per_t': R_total,
            }
    return out

# ─── Indirect LCA computation ───
def indirect_lca_per_country(slag_type, ci):
    """Reads BOF/BF/EAF dissolution result and computes indirect LCA at one CI."""
    fp = INDIRECT_INPUT_ROOT / f'{slag_type}/{slag_type}_Indirect_dissolution_result.xlsx'
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb['Sheet1']
    row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    slag_kg = row[0]
    diss_min = row[1]
    caco3 = row[3]
    prec_h = row[4]
    qfeed = row[5]

    # Crushing
    W_I = 0.0135
    DP = 1e-5
    D_sm = 0.0025
    E_crushing = 0.01 * W_I * (1.0/math.sqrt(DP) - 1.0/math.sqrt(D_sm)) * slag_kg
    # Stirring R1
    E_stir1 = (99.8121 / 1000) * (diss_min / 60)
    # Stirring R2
    Q_G = qfeed / 3600
    Re_L = (1050 * 1.5 * 0.6**2) / 1.2e-3
    Fr_L = (0.6 * 1.5**2) / 9.81
    Np = min(19.5 * Re_L**(-0.3), 24.0 * (Re_L * Fr_L)**(-1/3))
    P_0 = 1050 * 1.5**3 * 0.6**5 * Np
    term1 = (Q_G / (1.5 * 2))**(-0.25)
    term2 = ((1.5**2 * 0.6**4) / (9.81 * 0.12 * 2**(2/3)))**(-0.2)
    PG = P_0 * 0.1 * term1 * term2
    E_stir2 = PG / 1000 * prec_h
    # Fan
    delta_P = 1050 * 9.81 * 1
    P_in = 1e5
    P_out = P_in + delta_P
    gamma = 1.28
    eta = 0.85 * 0.95
    E_fan_W = (1/eta) * P_in * Q_G * gamma / (gamma - 1) * ((P_out/P_in)**(1-1/gamma) - 1)
    E_fan = E_fan_W / 1000 * prec_h
    # CO2 / DAC / BPMED
    CO2_stored = caco3 * 2000 * 44 / 1000
    E_DAC = (4552.583826 / 3.6) / 1000 * CO2_stored
    E_BPMED = 17328.25801 * (2000/2) / 3.6e6
    E_total = E_DAC + E_BPMED + E_stir1 + E_stir2 + E_fan + E_crushing

    conv = 1000 / slag_kg  # to per-tonne

    # Apply CI
    DAC_em = E_DAC * ci
    BPMED_em = E_BPMED * ci
    stir1_em = E_stir1 * ci
    stir2_em = E_stir2 * ci
    fan_em = E_fan * ci
    crush_em = E_crushing * ci
    total_em = E_total * ci
    net = total_em - CO2_stored
    net_pt = net * conv

    return {
        'DAC_0.1bar_kg': DAC_em * conv,
        'BPMED_kg': BPMED_em * conv,
        'Stirring_R1_kg': stir1_em * conv,
        'Stirring_R2_kg': stir2_em * conv,
        'Fan_kg': fan_em * conv,
        'Crushing_kg': crush_em * conv,
        'Total_emission_kg_per_t': total_em * conv,
        'CO2_stored_kg_per_t': CO2_stored * conv,
        'Net_emission_kg_per_t': net_pt,
    }

# ─── Build the master XLSX ───
print('Building Complete_LCA_OWID2025.xlsx...')

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='4F81BD')
SUBHEAD_FILL = PatternFill('solid', fgColor='DCE6F1')

def add_header_row(ws, row_idx, headers):
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=j, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Sheet 1: README
ws = wb.create_sheet('README')
ws.append(['Complete LCA dataset — steel slag carbonation, 37 countries, OWID 2025 grid CI'])
ws.append([])
ws.append(['Sheet', 'Description'])
sheets_meta = [
    ('OWID_2025_CI', '37 countries with electricity carbon intensities (kg CO2/kWh), Our World in Data 2025'),
    ('Direct_BOF', 'Gas-solid carbonation, BOF slag, 4 scenarios × 7 distances × 37 countries (1,036 rows)'),
    ('Direct_BF', 'Gas-solid carbonation, BF slag, same structure (1,036 rows)'),
    ('Direct_EAF', 'Gas-solid carbonation, EAF slag, same structure (1,036 rows)'),
    ('Indirect_BOF', 'Indirect aqueous carbonation, BOF slag, 37 countries (37 rows)'),
    ('Indirect_BF', 'Indirect aqueous carbonation, BF slag, 37 countries'),
    ('Indirect_EAF', 'Indirect aqueous carbonation, EAF slag, 37 countries'),
    ('Master_PerTonne', 'Best pathway per country × slag type (per-tonne basis, kg CO2/t slag)'),
    ('National_CDR_Potential', 'Best per-tonne × slag production = national annual CDR potential (kt CO2/yr)'),
]
for s, d in sheets_meta:
    ws.append([s, d])
ws.append([])
ws.append(['Methodology'])
ws.append(['  Direct (gas-solid): chamber + DAC (1 bar) + liquefaction + transport + vaporisation'])
ws.append(['  Indirect aqueous: dissolution (1M HCl) → precipitation (CaCO3) + BPMED reagent regen + DAC (0.1 bar)'])
ws.append(['  Functional unit: 1 tonne of slag carbonated to 90% theoretical CO2 uptake'])
ws.append(['  DAC Option 1: Carbon Engineering (wet scrubbing, 76.9% net atmospheric capture, NG-assisted)'])
ws.append(['  DAC Option 2: Climeworks (solid sorbent, 100% net atmospheric capture, all-electric)'])
ws.append(['  Negative Net emission = net CO2 removal from atmosphere (favourable)'])
for col, width in zip('AB', (28, 90)):
    ws.column_dimensions[col].width = width

# Sheet: OWID 2025 CI
ws = wb.create_sheet('OWID_2025_CI')
add_header_row(ws, 1, ['Country', 'CI (kg CO2/kWh)', 'CI rank (1=lowest)'])
sorted_c = sorted(OWID_2025.items(), key=lambda x: x[1])
for i, (c, ci) in enumerate(sorted_c, 1):
    ws.append([c, ci, i])
for col, w in zip('ABC', (20, 18, 18)):
    ws.column_dimensions[col].width = w

# Direct LCA sheets per slag type
direct_headers = [
    'Country', 'CI (kg/kWh)', 'Scenario', 'Distance (km)',
    'DAC process (kg/t)', 'Liquefaction (kg/t)', 'Vaporisation (kg/t)',
    'Transport (kg/t)', 'Crushing (kg/t)', 'Fan (kg/t)',
    'CO2 escape (kg/t)', 'Total emission (kg/t)', 'DAC capture (kg/t, credit)',
    'CO2 stored (kg/t)', 'Net emission (kg/t)',
]
sorted_countries = sorted(OWID_2025.keys(), key=lambda c: OWID_2025[c])

for slag in ['BOF', 'BF', 'EAF']:
    ws = wb.create_sheet(f'Direct_{slag}')
    add_header_row(ws, 1, direct_headers)
    slag_data = load_slag_lca_data(slag)
    for c in sorted_countries:
        ci = OWID_2025[c]
        for d in DISTANCES:
            res = direct_lca_per_country(slag_data, ci, d)
            for sname, comp in res.items():
                ws.append([
                    c, ci, sname, d,
                    comp['DAC_process_kg'], comp['Liquefaction_kg'], comp['Vaporisation_kg'],
                    comp['Transport_kg'], comp['Crushing_kg'], comp['Fan_kg'],
                    comp['CO2_escape_kg'], comp['Total_emission_kg'],
                    comp['DAC_capture_kg'], comp['CO2_stored_kg'],
                    comp['Net_emission_kg_per_t'],
                ])
    # Auto-width
    for col, w in zip('ABCDEFGHIJKLMNO', (18,12,15,12,18,18,18,15,15,12,15,18,20,15,18)):
        ws.column_dimensions[col].width = w
    print(f'  ✓ Direct_{slag}: {len(sorted_countries) * len(DISTANCES) * 4} rows')

# Indirect LCA sheets per slag type
indirect_headers = [
    'Country', 'CI (kg/kWh)',
    'DAC 0.1bar (kg/t)', 'BPMED (kg/t)', 'Stirring R1 (kg/t)', 'Stirring R2 (kg/t)',
    'Fan (kg/t)', 'Crushing (kg/t)',
    'Total emission (kg/t)', 'CO2 stored (kg/t)', 'Net emission (kg/t)',
]
for slag in ['BOF', 'BF', 'EAF']:
    ws = wb.create_sheet(f'Indirect_{slag}')
    add_header_row(ws, 1, indirect_headers)
    for c in sorted_countries:
        ci = OWID_2025[c]
        res = indirect_lca_per_country(slag, ci)
        ws.append([
            c, ci,
            res['DAC_0.1bar_kg'], res['BPMED_kg'], res['Stirring_R1_kg'], res['Stirring_R2_kg'],
            res['Fan_kg'], res['Crushing_kg'],
            res['Total_emission_kg_per_t'], res['CO2_stored_kg_per_t'], res['Net_emission_kg_per_t'],
        ])
    for col, w in zip('ABCDEFGHIJK', (18,12,18,15,18,18,12,15,18,15,18)):
        ws.column_dimensions[col].width = w
    print(f'  ✓ Indirect_{slag}: {len(sorted_countries)} rows')

# Master Per-Tonne summary
ws = wb.create_sheet('Master_PerTonne')
add_header_row(ws, 1, [
    'Country', 'CI (kg/kWh)', 'Slag type',
    'Net Indirect (kg/t)', 'Net GS_Opt1_Diesel (kg/t)', 'Net GS_Opt2_Diesel (kg/t)',
    'Net GS_Opt1_EV (kg/t)', 'Net GS_Opt2_EV (kg/t)',
    'Best pathway', 'Best Net (kg/t)', 'Regime',
])

# Load direct data for all slags (1st row of LCA, distance 0)
direct_data_by_slag = {s: load_slag_lca_data(s) for s in ['BOF', 'BF', 'EAF']}

def regime(ci):
    if ci > 0.445:
        return 'Deferred'
    elif ci < 0.192:
        return 'Indirect-favorable'
    else:
        return 'Gas-solid-favorable'

for c in sorted_countries:
    ci = OWID_2025[c]
    for slag in ['BOF', 'BF', 'EAF']:
        ind = indirect_lca_per_country(slag, ci)['Net_emission_kg_per_t']
        gs = direct_lca_per_country(direct_data_by_slag[slag], ci, 0)
        cand = [
            ('Indirect', ind),
            ('GS_Opt1_Diesel', gs['Diesel+Opt1']['Net_emission_kg_per_t']),
            ('GS_Opt2_Diesel', gs['Diesel+Opt2']['Net_emission_kg_per_t']),
            ('GS_Opt1_EV',     gs['EV+Opt1']['Net_emission_kg_per_t']),
            ('GS_Opt2_EV',     gs['EV+Opt2']['Net_emission_kg_per_t']),
        ]
        best_n, best_v = min(cand, key=lambda x: x[1])
        ws.append([
            c, ci, slag,
            ind,
            gs['Diesel+Opt1']['Net_emission_kg_per_t'], gs['Diesel+Opt2']['Net_emission_kg_per_t'],
            gs['EV+Opt1']['Net_emission_kg_per_t'],     gs['EV+Opt2']['Net_emission_kg_per_t'],
            best_n, best_v, regime(ci),
        ])

for col, w in zip('ABCDEFGHIJK', (18,12,10,18,22,22,22,22,15,15,20)):
    ws.column_dimensions[col].width = w
print(f'  ✓ Master_PerTonne: {len(sorted_countries) * 3} rows')

# National CDR potential
ws = wb.create_sheet('National_CDR_Potential')
add_header_row(ws, 1, [
    'Country', 'CI (kg/kWh)', 'Slag type', 'Slag production (Mt/yr)',
    'Best per-tonne Net (kg/t)', 'Best pathway',
    'National annual potential (kt CO2/yr)', 'Regime',
])
for c in sorted_countries:
    ci = OWID_2025[c]
    for slag in ['BOF', 'BF', 'EAF']:
        slag_mt = SLAG_PROD.get(c, {}).get(slag, 0)
        if slag_mt == 0:
            continue
        ind = indirect_lca_per_country(slag, ci)['Net_emission_kg_per_t']
        gs = direct_lca_per_country(direct_data_by_slag[slag], ci, 0)
        cand = [
            ('Indirect', ind),
            ('GS_Opt1_Diesel', gs['Diesel+Opt1']['Net_emission_kg_per_t']),
            ('GS_Opt2_Diesel', gs['Diesel+Opt2']['Net_emission_kg_per_t']),
            ('GS_Opt1_EV',     gs['EV+Opt1']['Net_emission_kg_per_t']),
            ('GS_Opt2_EV',     gs['EV+Opt2']['Net_emission_kg_per_t']),
        ]
        best_n, best_v = min(cand, key=lambda x: x[1])
        # National total: kg/t × Mt/yr = thousand_tonnes/yr (kt) [since 1 Mt = 1e6 t, × kg/t × 1e-3 = kt]
        national_kt = best_v * slag_mt * 1e-3
        # Wait check: best_v kg/t × slag_mt Mt = best_v * slag_mt * 1e6 kg / 1e6 = best_v * slag_mt Gg = best_v * slag_mt * 1e3 kg
        # Actually:  slag_mt (Mt/yr) × best_v (kg/t) = slag_mt × 1e6 t/yr × best_v kg/t = slag_mt * best_v * 1e6 kg/yr = slag_mt * best_v Gg/yr = slag_mt * best_v kt/yr
        national_kt = slag_mt * best_v
        ws.append([
            c, ci, slag, slag_mt,
            best_v, best_n,
            national_kt, regime(ci),
        ])
for col, w in zip('ABCDEFGH', (18,12,10,18,22,18,28,20)):
    ws.column_dimensions[col].width = w

# Save
wb.save(OUT_PATH)
print(f'\n✓ Complete_LCA_OWID2025.xlsx saved: {OUT_PATH}')
print(f'  Total: 1 README + 1 CI table + 3 Direct sheets + 3 Indirect sheets + 2 Master sheets')
import os
print(f'  Size: {os.path.getsize(OUT_PATH) / 1024:.1f} KB')
