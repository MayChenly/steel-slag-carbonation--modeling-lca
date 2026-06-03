"""Python port of gas-solid (direct carbonation) LCA MATLAB scripts.

Source: Project3_Steel slag/direct carbonation/results/LCApart/
  - BOF_lca_world.m / BF_lca_world.m / EAF_lca_world.m
  - SS_lca_world_0km.m / SS_lca_world_50km.m
  - direct_best_option_0km.m / direct_best_option_50km.m

Uses OWID 2025 grid carbon intensities (bug fix: removed hardcoded 0.2376).
"""
import math
from pathlib import Path
import openpyxl

DATA_ROOT = Path('/sessions/gallant-loving-edison/mnt/Project3_Steel slag/direct carbonation/results/LCApart')
OUT_ROOT = Path('/sessions/gallant-loving-edison/mnt/outputs/lca_python/gas_solid')
OUT_ROOT.mkdir(parents=True, exist_ok=True)

# OWID 2025 CI (kg CO2/kWh)
OWID_2025 = {
    'Austria': 0.11691, 'Belgium': 0.14982, 'Bulgaria': 0.27555, 'Croatia': 0.15850,
    'Czechia': 0.40146, 'Finland': 0.05747, 'France': 0.04144, 'Germany': 0.32965,
    'Greece': 0.31509, 'Hungary': 0.16302, 'Italy': 0.28478, 'Luxembourg': 0.12338,
    'Netherlands': 0.25356, 'Poland': 0.58860, 'Portugal': 0.12791, 'Romania': 0.25075,
    'Slovakia': 0.09485, 'Slovenia': 0.18329, 'Spain': 0.15360, 'Sweden': 0.03526,
    'Türkiye': 0.47474, 'United Kingdom': 0.21741, 'Russia': 0.44973,
    'Canada': 0.19072, 'Mexico': 0.47402, 'United States': 0.38440,
    'Argentina': 0.34600, 'Brazil': 0.10995, 'Chile': 0.28949,
    'Egypt': 0.56323, 'South Africa': 0.69929, 'Iran': 0.65953,
    'China': 0.52534, 'India': 0.67013, 'Japan': 0.47726,
    'South Korea': 0.41706, 'Australia': 0.52518,
}

# Distances (km) — from MATLAB: distance_one_way = 0:500:3000
DISTANCES = list(range(0, 3001, 500))  # 0, 500, 1000, 1500, 2000, 2500, 3000

# Fixed parameters
CO2_GAS_DAC_TONNE = 1.0       # ton CO2 product per DAC unit
TARE_WEIGHT = 0.4212          # ton
ROAD_EF = 0.112               # kg CO2 / ton·km (diesel truck)
EV_KWH_PER_KM = 0.921         # kWh/km EV truck
TOTAL_WEIGHT = CO2_GAS_DAC_TONNE + 2 * TARE_WEIGHT  # ton

# DAC Option 1 (Carbon Engineering, wet scrubbing, 1 bar)
O1_CO2_REMOVED_FROM_AIR = 0.769231  # tonne (per tonne CO2 produced)
O1_ELEC_PER_TONNE_CO2 = 152.8836    # kWh/tonne CO2 (electric only)
NATGAS_PER_TONNE_CO2 = 1121.796     # kWh/tonne (natural gas, not used in electric scope)

# DAC Option 2 (Climeworks, solid sorbent)
O2_CO2_REMOVED_FROM_AIR = 1.0
O2_ELEC_PER_TONNE_CO2 = 7220.0 / 3.6  # kWh/tonne CO2

# CO2 conditioning (per tonne CO2)
LIQUEFACTION_KWH = 89.0
VAPORISATION_KWH = 5.4

# ─── Load per-slag-type LCA input data ───
def load_slag_lca_data(slag_type):
    """Read e.g. BOF_lca_data.xlsx and return relevant fields."""
    fp = DATA_ROOT / f'{slag_type}_lca_data.xlsx'
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb.active
    # Find columns
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    def col(name):
        return headers.index(name) + 1

    rows = []
    for r in range(2, ws.max_row + 1):
        v = {h: ws.cell(r, col(h)).value for h in headers if h is not None}
        if v.get('SS_total_mass_kg') is None:
            break
        rows.append(v)
    return rows

# ─── LCA computation per country per distance ───
def compute_country_scenarios(slag_row, ci):
    """For one slag-row (one operating condition) at one country CI,
       compute the 4 LCA scenarios (diesel|EV × Opt1|Opt2) across all distances.
       Returns dict: {scenario_key: [(distance, R_total, components)…]}"""
    slag_mass = slag_row['SS_total_mass_kg']
    fan_MJ = slag_row['fan_energy_MJ']
    crush_kWh = slag_row['crushing_energy_kWh']
    CO2_supplied = slag_row['CO2_total_supplied_kg']
    CO2_stored = slag_row['CO2_total_stored_kg']

    conv = 1000.0 / slag_mass  # to per-tonne basis
    input_ratio = CO2_supplied / 1000.0 * conv

    # Per-tonne SCALAR (independent of country/distance)
    E_crushing = crush_kWh * ci * conv
    E_fan = fan_MJ / 3.6 * ci * conv
    E_escape = (CO2_supplied - CO2_stored) * conv

    # DAC processing (per tonne supplied → scaled to per-tonne slag via input_ratio)
    o1_E_DAC = O1_ELEC_PER_TONNE_CO2 * ci    # kg CO2 / t CO2 supplied
    o2_E_DAC = O2_ELEC_PER_TONNE_CO2 * ci
    E_liquefaction = LIQUEFACTION_KWH * ci
    E_vaporisation = VAPORISATION_KWH * ci

    # DAC capture credits (mass, CI-independent)
    o1_R_DAC = CO2_supplied * O1_CO2_REMOVED_FROM_AIR * conv
    o2_R_DAC = CO2_supplied * O2_CO2_REMOVED_FROM_AIR * conv
    CO2_stored_per_tonne = CO2_stored * conv

    results = {}
    for d in DISTANCES:
        # Diesel transport (return trip, so × 2)
        E_transp_diesel = TOTAL_WEIGHT * d * 2 * ROAD_EF   # kg CO2 per tonne CO2 supplied
        # EV transport
        E_transp_EV = d * 2 * EV_KWH_PER_KM * ci

        # Diesel + Opt 1
        o1_R_part1_d = o1_E_DAC + E_liquefaction + E_vaporisation + E_transp_diesel
        o1_R_total_d = o1_R_part1_d * input_ratio + E_crushing + E_fan + E_escape - o1_R_DAC

        # Diesel + Opt 2
        o2_R_part1_d = o2_E_DAC + E_liquefaction + E_vaporisation + E_transp_diesel
        o2_R_total_d = o2_R_part1_d * input_ratio + E_crushing + E_fan + E_escape - o2_R_DAC

        # EV + Opt 1
        e1_R_part1 = o1_E_DAC + E_liquefaction + E_vaporisation + E_transp_EV
        e1_R_total = e1_R_part1 * input_ratio + E_crushing + E_fan + E_escape - o1_R_DAC

        # EV + Opt 2
        e2_R_part1 = o2_E_DAC + E_liquefaction + E_vaporisation + E_transp_EV
        e2_R_total = e2_R_part1 * input_ratio + E_crushing + E_fan + E_escape - o2_R_DAC

        results.setdefault('diesel+opt1', []).append((d, o1_R_total_d))
        results.setdefault('diesel+opt2', []).append((d, o2_R_total_d))
        results.setdefault('EV+opt1', []).append((d, e1_R_total))
        results.setdefault('EV+opt2', []).append((d, e2_R_total))

    return {
        'scenarios': results,
        'CO2_stored_per_tonne': CO2_stored_per_tonne,
        'E_crushing': E_crushing,
        'E_fan': E_fan,
        'E_escape': E_escape,
    }

# ─── Run for each slag type ───
print('═' * 80)
print('Gas-solid pathway LCA — Python port, OWID 2025 CI')
print('═' * 80)

for slag_type in ['BOF', 'BF', 'EAF']:
    rows = load_slag_lca_data(slag_type)
    print(f'\n[{slag_type}] {len(rows)} LCA condition rows × 37 countries × 7 distances × 4 scenarios')
    # Sample first row of slag data
    r0 = rows[0]
    print(f'  Sample row: mass={r0["SS_total_mass_kg"]} kg, '
          f'crush={r0["crushing_energy_kWh"]:.2f} kWh, '
          f'fan={r0["fan_energy_MJ"]:.2f} MJ, '
          f'CO2_supplied={r0["CO2_total_supplied_kg"]:.2f} kg, '
          f'CO2_stored={r0["CO2_total_stored_kg"]:.2f} kg')

    # For each LCA scenario, build the all-countries × all-distances table
    sheets = {'diesel+opt1': [], 'diesel+opt2': [], 'EV+opt1': [], 'EV+opt2': []}
    header = ['Country', 'Distance (km)', 'R_total (kg CO2/t slag)', 'CI (kg/kWh)']
    for sname in sheets:
        sheets[sname].append(header)

    # Iterate over (slag_row × country)
    for slag_row in rows[:1]:  # only first row (UK case), as MATLAB uses first row
        for country in sorted(OWID_2025.keys()):
            ci = OWID_2025[country]
            res = compute_country_scenarios(slag_row, ci)
            for sname, vals in res['scenarios'].items():
                for d, total in vals:
                    sheets[sname].append([country, d, total, ci])

    # Save 4-sheet xlsx
    out_path = OUT_ROOT / f'{slag_type}_LCA_OWID2025.xlsx'
    wb = openpyxl.Workbook()
    first = True
    for sname, content in sheets.items():
        ws = wb.create_sheet(sname) if not first else wb.active
        if first:
            ws.title = sname
            first = False
        for row in content:
            ws.append(row)
    wb.save(out_path)
    print(f'  ✓ Saved {out_path.name}')

    # Show UK row for sanity
    print(f'\n  UK summary across 4 scenarios (at distance = 0 km):')
    slag_row = rows[0]
    uk_res = compute_country_scenarios(slag_row, OWID_2025['United Kingdom'])
    for sname, vals in uk_res['scenarios'].items():
        # value at distance 0
        d0_val = next(v for d, v in vals if d == 0)
        print(f'    {sname:<15s} (0 km): Net = {d0_val:+.2f} kg/t slag')

# ─── Best-option selector ───
print('\n')
print('═' * 80)
print('Best logistics × DAC configuration per country (BOF, 0 km transport)')
print('═' * 80)
rows = load_slag_lca_data('BOF')
slag_row = rows[0]
print(f'{"Country":<18s} {"CI":>8s} {"Best scenario":<18s} {"Net kg/t":>10s}')
print('-' * 60)
best_log = []
for country in sorted(OWID_2025.keys(), key=lambda c: OWID_2025[c]):
    ci = OWID_2025[country]
    res = compute_country_scenarios(slag_row, ci)
    # Find min across 4 scenarios at distance 0
    candidates = []
    for sname, vals in res['scenarios'].items():
        d0_val = next(v for d, v in vals if d == 0)
        candidates.append((sname, d0_val))
    candidates.sort(key=lambda x: x[1])
    best_s, best_net = candidates[0]
    best_log.append((country, ci, best_s, best_net))
    print(f'{country:<18s} {ci:>8.4f} {best_s:<18s} {best_net:>+10.2f}')

# Save best-option table
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'best_option_0km'
ws.append(['Country', 'CI', 'Best scenario', 'Net kg/t'])
for r in best_log:
    ws.append(list(r))
wb.save(OUT_ROOT / 'BOF_best_option_OWID2025.xlsx')
print(f'\n✓ Saved BOF_best_option_OWID2025.xlsx')
